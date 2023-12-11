from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from flask_socketio import SocketIO,emit
import time
import openai
import openpyxl 
import re
import tiktoken
import apikey
import time
import json
import langdetect
from dotenv import load_dotenv
import os
load_dotenv()
api_key = os.getenv('api_key')

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "http://localhost:5173"}})
socketio = SocketIO(app,cors_allowed_origins="http://localhost:5173")
@app.route('/')
def hello_world():
    return 'Connected to server'
@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files["file"]
    upload_location = 'upload/' + file.filename
    if 'file' not in request.files:
        return jsonify({'error': 'No file selected'}), 
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 
    file.save(upload_location)
    emit('uploaded', "Upload Success", broadcast=True, namespace='/')
    def num_tokens_from_messages(messages, model="gpt-3.5-turbo-0301"):
        """Returns the number of tokens used by a list of messages."""
        try:
            encoding = tiktoken.encoding_for_model(model)
        except KeyError:
            print("Warning: model not found. Using cl100k_base encoding.")
            encoding = tiktoken.get_encoding("cl100k_base")
        if model == "gpt-3.5-turbo":
            print("Warning: gpt-3.5-turbo may change over time. Returning num tokens assuming gpt-3.5-turbo-0301.")
            return num_tokens_from_messages(messages, model="gpt-3.5-turbo-0301")
        elif model == "gpt-4":
            print("Warning: gpt-4 may change over time. Returning num tokens assuming gpt-4-0314.")
            return num_tokens_from_messages(messages, model="gpt-4-0314")
        elif model == "gpt-3.5-turbo-0301":
            tokens_per_message = 4  # every message follows <|start|>{role/name}\n{content}<|end|>\n
            tokens_per_name = -1  # if there's a name, the role is omitted
        elif model == "gpt-4-0314":
            tokens_per_message = 3
            tokens_per_name = 1
        else:
            raise NotImplementedError(f"""num_tokens_from_messages() is not implemented for model {model}. See https://github.com/openai/openai-python/blob/main/chatml.md for information on how messages are converted to tokens.""")
        num_tokens = 0
        for message in messages:
            num_tokens += tokens_per_message
            for key, value in message.items():
                num_tokens += len(encoding.encode(value))
                if key == "name":
                    num_tokens += tokens_per_name
        num_tokens += 3  # every reply is primed with <|start|>assistant<|message|>
        return num_tokens
    system_prompt = """bạn là một phiên dịch viên tiếng nhật sang tiếng việt trong lĩnh vực IT. 
    hãy dịch những nội dung bên dưới sang tiếng việt theo phong cách dùng trong IT.
    nội dung cần dịch một câu sẽ theo định dạng bên dưới, tôi sẽ gửi nhiều câu trong một tin nhắn: 
    {{code}}:{{câu cần dịch}}

    trong đó câu cần dịch có thể nhiều dòng. dấu xuống dòng là @NEW_LINE_MARK@ và hãy giữ nguyên.

    Format tin nhắn của user sẽ như trong cặp dấu () sau:
    (
    INPUT
    ------
    A1: 更新履歴
    D6: 案件を選択

    GH25:-問い合わせる案件を選択してください。
    AH25:-画面ID

        作成者

    D7: 案件を選択

    )

    Format trả lời sẽ như bên dưới trong cặp dấu ():
    (
    RESULT
    ------
    A1: Lịch sử
    D6: Lựa chọn Anken
    GH25:-Hãy lựa chọn Anken contact
    AH25:-ID màn hình

        Người tạo

    D7: Lựa chọn Anken
    )
    hãy dịch toàn bộ tin nhắn  sang tiếng việt với cùng định dạng như trên, nếu bạn không thể dịch được thì hãy để trống nội dung phía sau dấu. Nếu nội dung cần dịch chứa tiếng việt hoặc tiếng anh thì cụm tiếng việt/ tiếng anh đó không cần dịch."""

    def translate(sheet):
        token = 0
        token_accumulator = 0
        openai.api_key = api_key
        
        new_translations = {}
        
        with open('dictJsonver1.json', 'r', encoding='utf-8') as json_file:
            dictionary = json.load(json_file)
        for row in sheet.iter_rows():
            row_str = ""
            for cell in row:
                cell_value = cell.value
                if cell_value in dictionary:
                    cell.value = dictionary[cell_value]
                if cell_value in dictionary.values():
                    continue
                if cell.value is not None and isinstance(cell.value, str) and not cell.value.isascii() and cell_value not in dictionary.keys():
                    row_str += """
                        {}:{}""".format(cell.coordinate, cell.value.replace("\n", " @NEW_LINE_MARK@ "))
                        
                    message_tokens = num_tokens_from_messages([{"role": "user", "content": cell.value}], model= "gpt-3.5-turbo-0301")
                    
                    if token_accumulator + message_tokens > 2000:
                        break
                    token_accumulator += message_tokens
            if row_str != "":
                messages=[
                        {
                            "role": "system", "content": system_prompt,
                        },
                        {
                            "role": "user", "content": """
                                INPUT
                                ------
                                {}
                            """.format(row_str),
                        },
                    ]
                message_tokens = num_tokens_from_messages(messages, model = "gpt-3.5-turbo-0301")
                token_accumulator += message_tokens 
                try:
                    response = openai.ChatCompletion.create(
                        model="gpt-3.5-turbo",
                        messages = messages
                    )
                except Exception as e:
                    print("Rate limit exzceeded. Please try again later.")
                    print(e)
                    emit('process', "Exceeded rate limit", broadcast=True, namespace='/')
                    print("30s until next request")
                    emit('process', "30s until next request", broadcast=True, namespace='/')
                    time.sleep(30)
                    print("continue")
                    continue
                print(token_accumulator)
                print(row_str)
                emit('process', "Translating"+ row_str, broadcast=True, namespace='/')
                if token_accumulator > 4000:
                    print("30s until next request")
                    emit('process', "30s until next request", broadcast=True, namespace='/')
                    time.sleep(30)
                    print("continue")
                    emit('process', "continue", broadcast=True, namespace='/')
                    token_accumulator = 0
                token += response['usage']['total_tokens']
                translated = response['choices'][0]['message']['content']

                regex = r"([a-z]+\d+):(.+)"
                matches = re.finditer(regex, translated, re.MULTILINE | re.IGNORECASE)
                for matchNum, match in enumerate(matches, start=1):
                    new_word = sheet[match.group(1)].value
                    sheet[match.group(1)].value = match.group(2).replace(" @NEW_LINE_MARK@ ", "\n")
                    # print(f"set {match.group(1)}={match.group(2)}")
                    cell_translation = match.group(2).replace(" @NEW_LINE_MARK@ ", "\n")
                    if new_word not in dictionary.keys():
                        new_translations[new_word] = cell_translation

        # Update the dictionary JSON file with new translations
        with open('dictJsonver1.json', 'r', encoding='utf-8') as json_file:
            dictionary = json.load(json_file)

        dictionary.update(new_translations)

        with open('dictJsonver1.json', 'w', encoding='utf-8') as json_file:
            json.dump(dictionary, json_file, ensure_ascii=False, indent=4)
        
        return token
    def process(input, output):
        total_tokens = 0
        workbook = openpyxl.load_workbook(input)
        sheet_names = workbook.sheetnames
        for sheet_name in sheet_names:
            original_sheet = workbook[sheet_name]
            total_tokens += translate(original_sheet)
            
            workbook.save(output)
        return total_tokens
    input = './upload/' + file.filename
    output = './translated/' + file.filename
    total_tokens = process(input, output)
    print("Total tokens: {}".format(total_tokens))
    emit('process', "Translating process done", broadcast=True, namespace='/')
    emit('complete', "Translation Complete", broadcast=True, namespace='/')
    return {'translated_file': file.filename}



@app.route('/download/<filename>', methods=['POST'])
def download_file(filename):
    translate_location = 'translated/' + filename
    return send_file(translate_location, as_attachment=True)


if __name__ == '__main__':
    app.run(port=5173)

